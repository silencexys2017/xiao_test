# -*- coding:utf-8 -*-
import sys
import os
import logging
import pymongo
import json
from kiliexpress.prd_kiliexpress import get_area_tree, get_pick_up_stations
from openpyxl import load_workbook

_DEFAULT_LOG_FILE = 'app.log'
_DEFAULT_CONFIG_FILE = '../../config.json'

CITY = ['Eldama Ravine', 'Mogotio', 'Chepalungu', 'Konoin', 'Sotik', 'Kabuchai', 'Kanduyi', 'Kimilili', 'Sirisia', 'Tongaren', 'Webuye', 'Butula', 'Funyula', 'Matayos', 'Nambale', 'Manyatta', 'Runyenjes', 'Balambala', 'Fafi', 'Ijara', 'Lagdera', 'Homa Bay Town', 'Karachuonyo', 'Kasipul', 'Mbita', 'Ndhiwa', 'Rangwe', 'Butere', 'Ikolomani', 'Khwisero', 'Likuyani', 'Lugari', 'Lurambi', 'Malava', 'Matungu', 'Mumias', 'Navakholo', 'Shinyalu', 'Ainamoi', 'Belgut', 'Bureti', 'Githunguri', 'Juja', 'Kabete', 'Kiambaa', 'Kiambu', 'Kikuyu', 'Lari', 'Limuru', 'Ruiru', 'Ganze', 'Kaloleni', 'Magarini', 'Malindi', 'Rabai', 'Gichugu', 'Mwea', 'Ndia', 'Bobasi', 'Bomachoge Borabu', 'Bomachoge Chache', 'Bonchari', 'South Mugirango', 'Muhoroni', 'Nyakach', 'Nyando', 'Seme', 'Kinango', 'Lunga Lunga', 'Matuga', 'Msambweni', 'Kathiani', 'Machakos Town', 'Masinga', 'Matungulu', 'Mavoko', 'Mwala', 'Yatta', 'Kaiti', 'Kilome', 'Mbooni', 'Banissa', 'Lafey', 'Laisamis', 'Moyale', 'Saku', 'Awendo', 'Nyatike', 'Rongo', 'Uriri', 'Changamwe', 'Jomvu', 'Kisauni', 'Likoni', 'Mvita', 'Nyali', 'Gatanga', 'Kandara', 'Kangema', 'Kigumo', 'Kiharu', 'Maragwa', 'Mathioya', 'Kamukunji', 'Kasarani', 'Kibra', 'Langata', 'Makadara', 'Mathare', 'Roysambu', 'Ruaraka', 'Starehe', 'Westlands', 'Bahati', 'Gilgil', 'Molo', 'Naivasha', 'Njoro', 'Rongai', 'Subukia', 'Aldai', 'Mosop', 'Nandi Hills', 'Tinderet', 'Emurua Dikirr', 'Kilgoris', 'Kitutu Masaba', 'Kinangop', 'Kipipiri', 'Ndaragwa', 'Ol Jorok', 'Ol Kalou', 'Kieni', 'Mathira', 'Mukurweini', 'Nyeri Town', 'Othaya', 'Tetu', 'Alego Usonga', 'Bondo', 'Gem', 'Rarieda', 'Ugenya', 'Ugunja', 'Mwatate', 'Taveta', 'Voi', 'Wundanyi', 'Bura', 'Galole', 'Garsen', 'Igambangombe', 'Maara', 'Tharaka', 'Endebess', 'Kiminini', 'Kwanza', 'Saboti', 'Loima', 'Ainabkoi', 'Kapseret', 'Moiben', 'Soy', 'Turbo', 'Emuhaya', 'Hamisi', 'Luanda', 'Sabatia', 'Eldas', 'Tarbaj', 'Kacheliba', 'Kapenguria', 'Sigor']
AREA = ['Eldama Ravine', 'Mogotio', 'Chepalungu', 'Sigor', 'Konoin', 'Sotik', 'Kabuchai', 'Kamukuywa', 'Tongaren', 'Nambale', 'Soy', 'Cherangany', 'Mwea', 'Balambala', 'Balambala', 'Bura', 'Fafi', 'Ijara', 'Ijara', 'Lagdera', 'Homa Bay Town', 'Karachuonyo', 'Kasipul', 'Mbita', 'Ndhiwa', 'Rangwe', 'Kitengela', 'Ngong', 'Olkeri', 'Ongata Rongai', 'Likuyani', 'Ainamoi', 'Litein', 'Chepseon', 'Kipkelion', 'Githunguri', 'Ikinu', 'Juja', 'Gitaru', 'Kabete', 'Uthiru', 'Tinganga', 'Kikuyu', 'Kinoo', 'Ganze', 'Kaloleni', 'Watamu', 'Magarini', 'Malindi', 'Rabai', 'Kerugoya', 'Ahero', 'Awasi', 'Yatta', 'Tharaka', 'Kinango', 'Lunga Lunga', 'Matuga', 'Msambweni', 'Nanyuki', 'Kathiani', 'Syokimau', 'Matuu', 'Kaiti', 'Kilome', 'Emali', 'Mbooni', 'Mbooni', 'Banissa', 'Banissa', 'Lafey', 'Lafey', 'Laisamis', 'Laisamis', 'Moyale', 'Saku', 'Awendo', 'Nyatike', 'Rongo', 'Uriri', 'Uriri', 'Changamwe', 'Miritini', 'Kongowea', 'Gatanga', 'Kilimani', 'Embakasi', 'Utawala', 'Imara Daima', 'Ruai', 'Karen', 'South C', 'Kahawa', 'Kahawa', 'Roysambu', 'Ngara', 'Pangani', 'Kangemi', 'Bahati', 'Molo', 'Njoro', 'Mosop', 'Subukia', 'Aldai', 'Kapsabet', 'Mosop', 'Nandi Hills', 'Nandi Hills', 'Tinderet', 'Tinderet', 'Emurua Dikirr', 'Kilgoris', 'Narok Town', 'Kitutu Masaba', 'Kinangop', 'Kipipiri', 'Kipipiri', 'Ndaragwa', 'Ol Jorok', 'Ol Kalou', 'Maralal', 'Alego Usonga', 'Bondo', 'Gem', 'Rarieda', 'Ugenya', 'Ukwala', 'Ugunja', 'Ugunja', 'Mwatate', 'Taveta', 'Kaloleni', 'Voi', 'Wundanyi', 'Bura', 'Galole', 'Garsen', 'Igambangombe', 'Endebess', 'Kiminini', 'Keiyo', 'Kwanza', 'Saboti', 'Loima', 'Lokichogio', 'Moiben', 'Soy', 'Emuhaya', 'Hamisi', 'Luanda', 'Luanda', 'Chavakali', 'Sabatia', 'Eldas', 'Eldas', 'Tarbaj', 'Tarbaj', 'Kapenguria']


def init_logging(filename):
    directory = os.path.dirname(filename)
    if directory != '' and not os.path.exists(directory):
        os.makedirs(directory)

    level = logging.INFO
    if os.environ.get('_DEBUG') == '1':
        level = logging.DEBUG
    fmt = '[%(asctime)s %(levelname)s | %(module)s %(funcName)s] %(message)s'
    logging.basicConfig(
        filename=filename, level=logging.DEBUG, format=fmt,
        datefmt="%Y-%M-%d %H:%M:%S")
    console = logging.StreamHandler()
    console.setLevel(level)
    console.setFormatter(logging.Formatter(fmt=fmt, datefmt="%H:%M:%S"))
    logging.getLogger().addHandler(console)


def load_config(filename, env):
    with open(filename, 'r') as f:
        config = json.loads(f.read())

    return config[env]


def get_db(config, env, database):
    uri = config['mongodb']['uri']
    client = pymongo.MongoClient(uri)
    db = '%s%s' % (env, database)
    return client[db]


def create_kili_express_address():
    bee_common_db.KiliCityApi.delete_many({})
    bee_common_db.KiliAreaApi.delete_many({})
    for region in get_area_tree()["data"]:
        if region.get("country") == "KE":
            for city in region.get("subAreaList"):
                bee_common_db.KiliCityApi.insert_one({
                    "id": city.get("areaId"),
                    "parentId": city.get("parentId"),
                    "name": city.get("areaName"),
                    "nameEn": city.get("areaNameEn"),
                    "country": city.get("country"),
                    "code": city.get("areaCode"),
                    "sort": city.get("areaSort"),
                    "status": city.get("status"),
                    "supportToDoor": city.get("supportToDoor")
                })

                for area in city.get("subAreaList") or []:
                    bee_common_db.KiliAreaApi.insert_one({
                        "id": city.get("areaId"),
                        "parentId": city.get("parentId"),
                        "name": area.get("areaName"),
                        "nameEn": area.get("areaNameEn"),
                        "country": area.get("country"),
                        "code": area.get("areaCode"),
                        "sort": area.get("areaSort"),
                        "status": area.get("status"),
                        "supportToDoor": area.get("supportToDoor")
                    })


def update_speedaf_address():
    region_id = bee_common_db.Region.find_one({"code": "KE"})["_id"]
    name_map = {
        "BUNGO MA": "Bungoma",
        "Elgeyo  Marakwet": "Elgeyo-Marakwet",
        "Murang’a": "Muranga",
        "HOMABA Y": "Homa Bay",
        "Trans Nzoia": "Trans-Nzoia",
        "Taita Taveta": "Taita-Taveta",
        "Tharaka   Nithi": "Tharaka-Nithi",
        }
    for state in bee_common_db.State.find({"regionId": region_id}):
        if name_map.get(state["name"]):
            name = name_map.get(state["name"])
        else:
            if "-" in state["name"]:
                split_names = [s.capitalize() for s in state["name"].split("-")]
                name = "-".join(split_names)
            else:
                split_names = [s.capitalize() for s in state["name"].split(" ")]
                name = " ".join(split_names)
        if state["name"] != name:
            bee_common_db.State.update_one({"_id": state["_id"]},
                                           {"$set": {"name": name}})

    for city in bee_common_db.City.find({"regionId": region_id}):
        name = city["name"]
        # name[name.find("("):name.find(")") + 1]
        if "(New)" in name:
            name = name.replace('(New)', '')
        if "( New)" in name:
            name = name.replace('( New)', '')
        if "(new)" in name:
            name = name.replace('(new)', '')
        if "( new)" in name:
            name = name.replace('( new)', '')
        if '"' in name:
            name = name.replace('"', '')
        if "'" in name:
            name = name.replace("'", "")
        if "’" in name:
            name = name.replace("’", "")
        if "." in name:
            name = name.replace(".", "")
        name = name.strip()
        split_names = [s.capitalize() for s in name.split(" ")]
        name = " ".join(split_names)
        if city["name"] != name:
            bee_common_db.City.update_one({"_id": city["_id"]},
                                          {"$set": {"name": name}})
    for area in bee_common_db.Area.find({"regionId": region_id}):
        name = area["name"]
        if "(New)" in name:
            name = name.replace('(New)', '')
        if "( New)" in name:
            name = name.replace('( New)', '')
        if "(new)" in name:
            name = name.replace('(new)', '')
        if "( new)" in name:
            name = name.replace('( new)', '')
        if '"' in name:
            name = name.replace('"', '')
        if "'" in name:
            name = name.replace("'", "")
        if "’" in name:
            name = name.replace("’", "")
        if "." in name:
            name = name.replace(".", "")
        if "/" in name:
            name = name.replace("/", " ")
        if "–" in name:
            name = name.replace("–", " ")
        if "," in name:
            name = name.replace(",", " ")

        name = name.strip()
        split_names = [s.capitalize() for s in name.split(" ")]
        split_names = list(filter(None, split_names))
        name = " ".join(split_names)
        if area["name"] != name:
            bee_common_db.Area.update_one({"_id": area["_id"]},
                                          {"$set": {"name": name}})


def update_s2b_address():
    name_map = {
        "BUNGO MA": "Bungoma",
        "Elgeyo  Marakwet": "Elgeyo-Marakwet",
        "Murang’a": "Muranga",
        "HOMABA Y": "Homa Bay",
        "Trans Nzoia": "Trans-Nzoia",
        "Taita Taveta": "Taita-Taveta",
        "Tharaka   Nithi": "Tharaka-Nithi",
        }
    for state in quark_common_db.State.find({"regionCode": "KE"}):
        if name_map.get(state["name"]):
            name = name_map.get(state["name"])
        else:
            if "-" in state["name"]:
                split_names = [s.capitalize() for s in state["name"].split("-")]
                name = "-".join(split_names)
            else:
                split_names = [s.capitalize() for s in state["name"].split(" ")]
                name = " ".join(split_names)
        if state["name"] != name:
            quark_common_db.State.update_one({"_id": state["_id"]},
                                           {"$set": {"name": name}})

    for city in quark_common_db.City.find({"regionCode": "KE"}):
        name = city["name"]
        # name[name.find("("):name.find(")") + 1]
        if "(New)" in name:
            name = name.replace('(New)', '')
        if "( New)" in name:
            name = name.replace('( New)', '')
        if "(new)" in name:
            name = name.replace('(new)', '')
        if "( new)" in name:
            name = name.replace('( new)', '')
        if '"' in name:
            name = name.replace('"', '')
        if "'" in name:
            name = name.replace("'", "")
        if "’" in name:
            name = name.replace("’", "")
        if "." in name:
            name = name.replace(".", "")
        name = name.strip()
        split_names = [s.capitalize() for s in name.split(" ")]
        name = " ".join(split_names)
        if city["name"] != name:
            quark_common_db.City.update_one({"_id": city["_id"]},
                                          {"$set": {"name": name}})
    for area in quark_common_db.Area.find({"regionCode": "KE"}):
        name = area["name"]
        if "(New)" in name:
            name = name.replace('(New)', '')
        if "( New)" in name:
            name = name.replace('( New)', '')
        if "(new)" in name:
            name = name.replace('(new)', '')
        if "( new)" in name:
            name = name.replace('( new)', '')
        if '"' in name:
            name = name.replace('"', '')
        if "'" in name:
            name = name.replace("'", "")
        if "’" in name:
            name = name.replace("’", "")
        if "." in name:
            name = name.replace(".", "")
        if "/" in name:
            name = name.replace("/", " ")
        if "–" in name:
            name = name.replace("–", " ")
        if "," in name:
            name = name.replace(",", " ")

        name = name.strip()
        split_names = [s.capitalize() for s in name.split(" ")]
        split_names = list(filter(None, split_names))
        name = " ".join(split_names)
        if area["name"] != name:
            quark_common_db.Area.update_one({"_id": area["_id"]},
                                          {"$set": {"name": name}})

    for it in quark_auth_db.Address.find({"regionCode": "KE"}):
        state = quark_common_db.State.find_one({"_id": it["stateId"]})["name"]
        city = quark_common_db.City.find_one({"_id": it["cityId"]})["name"]
        area = quark_common_db.Area.find_one({"_id": it["areaId"]})["name"]
        quark_auth_db.Address.update_one(
            {"_id": it["_id"]}, {"$set": {
                "state": state, "city": city, "area": area}})


def create_kili_address():
    bee_common_db.KiliCity.delete_many({})
    bee_common_db.KiliArea.delete_many({})
    wb_obj = load_workbook("./kiliexpress/kili_base_area.xlsx")
    sheet = wb_obj.active
    for row in list(sheet.iter_rows(min_row=2)):
        area_id = row[0].value
        parent_id = row[1].value
        level = row[2].value
        region_code = row[3].value
        name = row[5].value
        code = row[6].value
        support_to_door = row[10].value
        status = row[11].value
        if region_code == "KE":
            if level == 1:
                bee_common_db.KiliCity.insert_one({
                    "id": area_id,
                    "parentId": parent_id,
                    "name": name,
                    "country": region_code,
                    "code": code,
                    "status": status,
                    "supportToDoor": support_to_door
                })
            elif level == 2:
                bee_common_db.KiliArea.insert_one({
                    "id": area_id,
                    "parentId": parent_id,
                    "name": name,
                    "country": region_code,
                    "code": code,
                    "status": status,
                    "supportToDoor": support_to_door
                })


def diff_kili_and_self_address():
    region_id = bee_common_db.Region.find_one({"code": "KE"})["_id"]
    for state in bee_common_db.KiliCity.find():
        if not bee_common_db.State.find_one(
                {"regionId": region_id, "name": state["name"]}):
            print("not found county = %r" % state)

    match_city = []
    for city in bee_common_db.City.find({"regionId": region_id}):
        res = bee_common_db.KiliArea.find_one(
                {"name": {"$regex": city["name"], "$options": 'i'}})
        if res:
            match_city.append(city["name"])
            if city["name"] not in CITY:
                print("self=%s,kili=%s" % (city["name"], res["name"]))
    print("-------match-city--------count=%s" % len(match_city))
    print(match_city)
    match_area = []
    for area in bee_common_db.Area.find({"regionId": region_id}):
        res = bee_common_db.KiliArea.find_one(
                {"name": {"$regex": area["name"], "$options": 'i'}})
        if res:
            match_area.append(area["name"])
            if area["name"] not in AREA:
                print("self=%s,kili=%s" % (area["name"], res["name"]))
    print("-------match-area--------count=%s" % len(match_area))
    print(match_area)


def get_match_address():
    wb_obj = load_workbook("./kiliexpress/ke_af_map.xlsx")
    sheet = wb_obj.active
    match_city = []
    for cell in sheet['D']:
        if cell.value:
            split_names = [s.capitalize() for s in cell.value.split(" ")]
            name = " ".join(split_names)
            match_city.append(name)
    print("-------match-city--------count=%s" % len(match_city))
    print(match_city)
    match_area = []
    for cell in sheet['E']:
        if cell.value:
            split_names = [s.capitalize() for s in cell.value.split(" ")]
            name = " ".join(split_names)
            match_area.append(name)
    print("-------match-area--------count=%s" % len(match_area))
    print(match_area)

    error_city = []
    for it in match_city:
        if it not in CITY:
            error_city.append(it)
    print("-------error-city--------count=%s" % len(error_city))
    print(error_city)
    error_area = []
    for it in match_area:
        if it not in AREA:
            error_area.append(it)
    print("-------error-area--------count=%s" % len(error_area))
    print(error_area)


def create_xed_match_address():
    bee_logistics_db.KiliAddressMapping.delete_many({})
    region_id = bee_common_db.Region.find_one({"code": "KE"})["_id"]
    k_city = {it["id"]: it["name"] for it in bee_common_db.KiliCity.find()}
    for area in bee_common_db.KiliArea.find().sort([("parentId", 1)]):
        county = k_city[area["parentId"]]
        if county == "Pokot West":
            state_name = "West Pokot"
        else:
            state_name = county
        state_id = bee_common_db.State.find_one(
            {"regionId": region_id, "name": state_name})["_id"]
        city_name = "Makueni" if area["name"] == "Makueni Town" else area["name"]
        city = bee_common_db.City.find_one(
            {"name": city_name, "stateId": state_id})
        city_id = city["_id"] if city else None
        area_ids = bee_common_db.Area.distinct(
            "_id", {"regionId": region_id, "cityId": city_id})
        self_area = bee_common_db.Area.find_one(
            {"regionId": region_id, "cityId": city_id})
        if self_area:
            area_ids.append(self_area["_id"])
            area_ids = list(set(area_ids))
        bee_logistics_db.KiliAddressMapping.insert_one(
            {
                "areaId": area["id"],
                "areaCode": area["code"],
                "county": county,
                "town": area["name"],
                "selfStateId": state_id,
                "selfCityId": city_id,
                "selfAreaIds": area_ids,
                "isSupportToDoor": area["supportToDoor"],
                "state": area["status"],
            }
        )


def diff_excel_data():
    wb_obj = load_workbook("./kiliexpress/ke_af_map.xlsx")
    sheet = wb_obj.active
    match_town = []
    for cell in sheet['B']:
        if cell.value:
            match_town.append(cell.value)
    for it in bee_common_db.KiliArea.find():
        if it["name"] not in match_town:
            print(it["name"])


def create_s2b_pickup_station():
    quark_common_db.PickupStation.delete_many({})
    pickup_station_id = 0
    for it in bee_logistics_db.KiliAddressMapping.find().sort([("areaId", 1)]):
        if it["county"] == "Pokot West":
            state_name = "West Pokot"
        else:
            state_name = it["county"]
        state_id = quark_common_db.State.find_one(
            {"regionCode": "KE", "name": state_name})["_id"]

        city_name = "Makueni" if it["town"] == "Makueni Town" else it["town"]
        city = quark_common_db.City.find_one(
            {"name": city_name, "stateId": state_id})
        city_id = city["_id"] if city else None
        area_ids = quark_common_db.Area.distinct(
            "_id", {"regionCode": "KE", "cityId": city_id})
        self_area = quark_common_db.Area.find_one(
            {"regionCode": "KE", "cityId": city_id})
        if self_area:
            area_ids.append(self_area["_id"])
            area_ids = list(set(area_ids))
        for station in get_pick_up_stations(it["areaCode"]).get("data") or []:
            pickup_station_id += 1
            quark_common_db.PickupStation.insert_one(
                {
                    "_id": pickup_station_id,
                    "sourceStationId": station["id"],
                    "name": station["stationName"],
                    "address": station["address"],
                    "isSupportBigPackage": station.get("isSupportBigPackage"),
                    "areaId": station["areaId"],
                    "areaCode": station["areaCode"],
                    "businessType": station["businessType"],
                    "pickingCode": station["pickingCode"],
                    "businessTime": station["businessTime"],
                    'latitude': station["latitude"],
                    'longitude': station["longitude"],
                    "regionCode": "KE",
                    "county": it["county"],
                    "town": it["town"],
                    "selfStateId": state_id,
                    "selfCityId": city_id,
                    "selfAreaIds": area_ids
                })


def add_related_mongo_data():
    _id = 0
    for it in bee_logistics_db.TransportCorridor.find().sort(
            [("_id", -1)]).limit(1):
        _id = it["_id"] + 1
    bee_logistics_db.TransportCorridor.insert_one({
        "_id": _id,
        "index": _id,
        "providerId": "KiliExpress",
        "providerLocation": "易递-KiliExpress",
        "providerName": "KiliExpress",
        "contactName": "Harunur Roshid",
        "contactPhone": "01840566762",
        "type": 2,
        "terminalRegionIds": [
                8
        ],
        "state": 1,
        "targetRegionIds": [
                8
        ],
        "supportAssortments": [
                1, 2, 3, 4, 5, 6
        ],
        "terminalType": {
                "1": 1
        }
    })
    quark_auth_db.Address.update_many(
        {"addressType": {"$exists": False}},
        {"$set": {"addressType": 1}})

    bee_logistics_db.LogisticsOrder.update_many(
        {"isPickup": {"$exists": False}},
        {"$set": {"isPickup": False, "recipient.networkId": None}})

    per_id = 0
    for it in quark_admin_db.Permission.find().sort([("_id", -1)]).limit(1):
        per_id = it["_id"] + 1
    quark_admin_db.Permission.insert_one(
        {
            "_id": per_id,
            "name": "自提点设置",
            "code": "pickupStationPostage",
            "codePath": "pickupStationPostage",
            "idPath": str(per_id),
            "type": 1,
            "parentId": None
        })


if __name__ == "__main__":
    usage = 'python3 Sxx.py prd|dev|test'
    init_logging(_DEFAULT_LOG_FILE)
    if len(sys.argv) < 2:
        logging.error(usage)
        sys.exit(-1)

    env = sys.argv[1]
    config = load_config(_DEFAULT_CONFIG_FILE, env)
    bee_common_db = get_db(config, env, "BeeCommon")
    quark_common_db = get_db(config, env, "QuarkCommon")
    quark_auth_db = get_db(config, env, "QuarkAuth")
    quark_admin_db = get_db(config, env, "QuarkAdmin")
    bee_logistics_db = get_db(config, env, "BeeLogistics")

    # create_kili_express_address()

    create_kili_address()
    update_speedaf_address()
    update_s2b_address()
    # # diff_kili_and_self_address()
    create_xed_match_address()
    add_related_mongo_data()
    create_s2b_pickup_station()

    print("---------------------------success-------------------------------")

