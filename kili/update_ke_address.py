# -*- coding:utf-8 -*-
import sys
import os
import logging
import pymongo
import json
from openpyxl import load_workbook
from datetime import timedelta

_DEFAULT_CONFIG_FILE = '../config.json'

X_DB_URL = {
    "dev": "mongodb://pf_dev_dbo:Bp2Q5j3Jb2cmQvn8L4kW@mongodb-paas-service/admin?replicaSet=rs0",
	"test": "mongodb://pf_test_dbo:3f4k8aDHeQJBKmd3z7c9@159.138.90.30:30734/admin",
	# "prd": "mongodb://pf_prd_dbo:m2w9ZNZP4gUsx9b2hu6L@159.138.90.30:30734/admin?authSource=admin",
    "prd": "mongodb://pf_prd_dbo:m2w9ZNZP4gUsx9b2hu6L@mongodb-paas-service/admin?replicaSet=rs0"
}

K_DB_URL = {
    "dev": "mongodb://root:KB5NF1T0aP@mongodb-headless.os:27017/admin?replicaSet=rs0",
	"test": "mongodb://root:IdrCgVpHzv@mongo-mongodb-headless.os:27017/admin?replicaSet=rs0&retrywrites=false",
    "prd": "mongodb://sk2:Gmbz8i63mtswuMKz8Lk92zNM7UxwO5Txz1@ebes-db.cmxtf8qiglae.eu-central-1.docdb.amazonaws.com:27017/admin?replicaSet=rs0&retrywrites=false"
}

UTC_FORMAT = '%Y-%m-%dT%H:%M:%S.%fZ'
LOCAL_FORMAT = '%Y-%m-%d %H:%M:%S'


def init_logging(filename):
    directory = os.path.dirname(filename)
    if directory != '' and not os.path.exists(directory):
        os.makedirs(directory)

    level = logging.INFO
    if os.environ.get('_DEBUG') == '1':
        level = logging.DEBUG
    fmt = '[%(asctime)s %(levelname)s | %(module)s %(funcName)s] %(message)s'
    logging.basicConfig(
        filename=filename, level=logging.DEBUG, format=fmt,
        datefmt="%Y-%M-%d %H:%M:%S")
    console = logging.StreamHandler()
    console.setLevel(level)
    console.setFormatter(logging.Formatter(fmt=fmt, datefmt="%H:%M:%S"))
    logging.getLogger().addHandler(console)


def load_config(filename, env):
    with open(filename, 'r') as f:
        config = json.loads(f.read())

    return config[env]


def get_db(uri, env, database):
    client = pymongo.MongoClient(uri[env], retryWrites=False)
    db = '%s%s' % (env, database)
    return client[db]


def utc2local(utc_datetime):
    if not utc_datetime:
        return utc_datetime
    # utc_dt = datetime.strptime(utcstr, UTC_FORMAT)
    # local: 东6区
    lo_dt = utc_datetime + timedelta(hours=6)
    return lo_dt.strftime(LOCAL_FORMAT)


def update_excel_data(wb_obj, is_update_xed=False):
    # 使用指定工作表
    sheet = wb_obj.active  # 当前激活的工作表

    for row in list(sheet.iter_rows(min_row=2)):
        county, sub_county, word = row[2].value, row[3].value, row[4].value
        cod = row[6].value
        support_cod = False if cod in ["FALSE", False] else True
        if is_update_xed:
            area_1 = bee_common_db.Areas.find_one({"deep": 1, "name": county})
            if not area_1:
                print(county, sub_county, word, "----------county")
                continue
            area_2 = bee_common_db.Areas.find_one(
                {"deep": 2, "parentId": area_1["_id"], "name": sub_county})
            if not area_2:
                print(county, sub_county, word, "-----------subCounty")
                continue
            area_3 = bee_common_db.Areas.find_one(
                {"deep": 3, "parentId": area_2["_id"], "name": word})
            if area_3:
                bee_common_db.Areas.update_one(
                    {"_id": area_3["_id"]},
                    {"$set": {"supportCod": support_cod}})
        else:
            area_1 = common_db.Areas.find_one({"deep": 1, "name": county})
            if not area_1:
                print(county, sub_county, word, "----------county")
                continue
            area_2 = common_db.Areas.find_one(
                {"deep": 2, "parentId": area_1["_id"], "name": sub_county})
            if not area_2:
                print(county, sub_county, word, "-----------subCounty")
                continue
            area_3 = common_db.Areas.find_one(
                {"deep": 3, "parentId": area_2["_id"], "name": word})
            if area_3:
                common_db.Areas.update_one(
                    {"_id": area_3["_id"]},
                    {"$set": {"supportCod": support_cod}})


def update_pickup_station_support_cod(is_update_xed=False):
    stations = [718865213, 718865219, 718864953, 718864751, 36, 22, 718865222]
    if is_update_xed:
        bee_common_db.PickupStation.update_many(
            {"sourceStationId": {"$in": stations}},
            {"$set": {"isSupportCOD": True}})
    else:
        common_db.PickupStation.update_many(
            {"sourceStationId": {"$in": stations}},
            {"$set": {"isSupportCOD": True}})


def pickup_stations_update():
    common_db.PickupStation.delete_many({})
    for it in bee_common_db.PickupStation.find({}):
        if not it.get("phone"):
            it["phone"] = "132434342"
        common_db.PickupStation.insert_one(it)


if __name__ == "__main__":
    usage = 'python3 Sxx.py prd|dev|test'

    init_logging("xed-order-export.log")
    if len(sys.argv) < 2:
        logging.error(usage)
        sys.exit(-1)

    env, xed = sys.argv[1], sys.argv[2]
    if env == "prd":
        url = X_DB_URL
    else:
        url = K_DB_URL
    bee_common_db = get_db(url, env, "BeeCommon")
    common_db = get_db(K_DB_URL, env, "Common")
    FILE_NAME = "kenya_address.xlsx"
    if xed in [1, "1"]:
        is_update_xed = True
    else:
        is_update_xed = False
    # wb_obj = load_workbook(FILE_NAME)
    #
    # update_excel_data(wb_obj, is_update_xed)
    # update_pickup_station_support_cod(is_update_xed)
    # # 保存文件，也可覆盖原文件
    # wb_obj.save(FILE_NAME)

    pickup_stations_update()


