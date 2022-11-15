# -*- coding:utf-8 -*-
import sys
import os
import logging
import pymongo
import json
from datetime import datetime, timedelta
from openpyxl import load_workbook


_DEFAULT_CONFIG_FILE = '../kili_config.json'
_SOUTHX_CONFIG_FILE = '../config.json'
DATA_TIME_FORMAT = "%Y-%m-%dT%H:%M:%S.%fZ"

REGION_MATCH = {1: "Bangladesh", 3: "Myanmar"}
map_pay_method = {
    1: 'Online',
    2: 'COD',
    3: 'Pre'
}

X_DB_URL = {
    "dev": "mongodb://pf_dev_dbo:Bp2Q5j3Jb2cmQvn8L4kW@mongodb-paas-service/admin?replicaSet=rs0",
	"test": "mongodb://pf_test_dbo:3f4k8aDHeQJBKmd3z7c9@159.138.90.30:30734/admin",
	# "prd": "mongodb://pf_prd_dbo:m2w9ZNZP4gUsx9b2hu6L@159.138.90.30:30734/admin",
    "prd": "mongodb://pf_prd_dbo:m2w9ZNZP4gUsx9b2hu6L@mongodb-paas-service/admin?replicaSet=rs0"
}

K_DB_URL = {
    "dev": "mongodb://root:ty8NaMSWoM@10.0.1.103:32346/admin?readPreference=primaryPreferred",
	"test": "mongodb://root:ty8NaMSWoM@10.0.1.103:32346/admin",
	"prd": "mongodb://sk2:Qmvz84mtswuMJz8Kk90zBM7UbdW@ebes-db.cluster-cvxmlpsy4xpw.eu-central-1.docdb.amazonaws.com:27017/admin?authSource=admin"
}


UTC_FORMAT = '%Y-%m-%dT%H:%M:%S.%fZ'
LOCAL_FORMAT = '%Y-%m-%d %H:%M:%S'


def init_logging(filename):
    directory = os.path.dirname(filename)
    if directory != '' and not os.path.exists(directory):
        os.makedirs(directory)

    level = logging.INFO
    if os.environ.get('_DEBUG') == '1':
        level = logging.DEBUG
    fmt = '[%(asctime)s %(levelname)s | %(module)s %(funcName)s] %(message)s'
    logging.basicConfig(
        filename=filename, level=logging.DEBUG, format=fmt,
        datefmt="%Y-%M-%d %H:%M:%S")
    console = logging.StreamHandler()
    console.setLevel(level)
    console.setFormatter(logging.Formatter(fmt=fmt, datefmt="%H:%M:%S"))
    logging.getLogger().addHandler(console)


def load_config(filename, env):
    with open(filename, 'r') as f:
        config = json.loads(f.read())

    return config[env]


def get_db(uri, env, database):
    client = pymongo.MongoClient(uri[env])
    db = '%s%s' % (env, database)
    return client[db]


def utc2local(utc_datetime):
    if not utc_datetime:
        return utc_datetime
    # utc_dt = datetime.strptime(utcstr, UTC_FORMAT)
    # local: 东6区
    lo_dt = utc_datetime + timedelta(hours=6)
    return lo_dt.strftime(LOCAL_FORMAT)


def get_deep_start_id(deep):
    if deep in [0, 1, 2, 3]:
        return int(1000 * 10 ** ((deep - 1) * 3))
    elif deep > 3:
        return int(1000 * 10 ** (1+deep))


def convert_time(date_time_str):
    if date_time_str:
        return datetime.strptime(date_time_str, DATA_TIME_FORMAT)
    return None


def convert_time_str(date_time):
    if date_time:
        return date_time.strftime(DATA_TIME_FORMAT)
    return None


def strip_string_name(name):
    if not name:
        return
    name = name.strip()
    split_names = [s.capitalize() if "-" not in s else s for s in name.split(" ")]
    split_names = list(filter(None, split_names))
    name = " ".join(split_names)
    return name


def import_ke_address_from_excel(file_route):
    wb_obj = load_workbook(file_route, data_only=True)
    ks_sheet = wb_obj.get_sheet_by_name("KS&ward")
    total_count = 0
    for row in list(ks_sheet.iter_rows(min_row=3)):
        total_count += 1
        county, sub_county, area = strip_string_name(
            row[0].value), strip_string_name(row[1].value), strip_string_name(
            row[2].value)
        shop_id = strip_string_name(row[4].value)
        if not shop_id or not shop_id.isdigit():
            continue
        shop_id = int(shop_id)
        # address, shop_id, shop_name = row[3].value, row[4].value, row[5].value
        res_county = bee_common_db.Areas.find_one({"deep": 1, "name": county})
        if not res_county:
            continue
        res_sub_county = bee_common_db.Areas.find_one(
            {"deep": 2, "name": sub_county})
        if not res_sub_county:
            continue
        res_area = bee_common_db.Areas.find_one(
            {"deep": 3, "name": area})
        if not res_area:
            continue
        station = bee_common_db.PickupStation.find_one_and_update(
            {"sourceStationId": shop_id},
            {"$addToSet": {"leafAreaIds": res_area["_id"]}})
        if not station:
            continue
        bee_common_db.Areas.update_one(
            {"_id": res_area["_id"]},
            {"$addToSet": {"pickupIds": shop_id}})
        print("success area=%s,line=%s" % (area, total_count))
    print("loop count=%s" % total_count)


if __name__ == "__main__":
    usage = 'python3 Sxx.py prd|dev|test'
    init_logging("release_ready.log")
    if len(sys.argv) < 2:
        logging.error(usage)
        sys.exit(-1)

    env = sys.argv[1]
    # config = load_config(_DEFAULT_CONFIG_FILE, env)
    # southx_config = load_config(_SOUTHX_CONFIG_FILE, env)

    bee_common_db = get_db(X_DB_URL, env, "BeeCommon")
    bee_logistics_db = get_db(X_DB_URL, env, "BeeLogistics")

    import_ke_address_from_excel("./address Update V6.xlsx")

    ki_li_data = {"stations": [], "areas": []}
    for it in bee_common_db.PickupStation.find().sort([("_id", 1)]):
        if it.get("lastUpdatedTime"):
            it["lastUpdatedTime"] = convert_time_str(it["lastUpdatedTime"])
        ki_li_data["stations"].append(it)

    for it in bee_common_db.Areas.find(
            {"regionCode": {"$in": ["KE", "CN"]}}).sort([("deep", 1)]):
        it["lastUpdatedTime"] = convert_time_str(it["lastUpdatedTime"])
        ki_li_data["areas"].append(it)

    with open("ki_li_data.json", 'w', encoding="utf-8") as write_f:
        write_f.write(json.dumps(ki_li_data, indent=4, ensure_ascii=False))


