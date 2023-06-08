# -*- coding:utf-8 -*-
import sys
import os
import logging
import pymongo
import json
from datetime import datetime, timedelta
from openpyxl import load_workbook
import xlsxwriter
from prd_kiliexpress import sign_in, get_area_tree


_DEFAULT_CONFIG_FILE = '../kili_config.json'
_SOUTHX_CONFIG_FILE = '../config.json'
DATA_TIME_FORMAT = "%Y-%m-%dT%H:%M:%S.%fZ"

REGION_MATCH = {1: "Bangladesh", 3: "Myanmar"}
map_pay_method = {
    1: 'Online',
    2: 'COD',
    3: 'Pre'
}


K_DB_URL = {
    "dev": "mongodb://root:KB5NF1T0aP@mongodb-headless.os:27017/admin?replicaSet=rs0",
	"test": "mongodb://root:IdrCgVpHzv@mongo-mongodb-headless.os:27017/admin?replicaSet=rs0&retrywrites=false",
    "prd": ""
}


UTC_FORMAT = '%Y-%m-%dT%H:%M:%S.%fZ'
LOCAL_FORMAT = '%Y-%m-%d %H:%M:%S'


def init_logging(filename):
    directory = os.path.dirname(filename)
    if directory != '' and not os.path.exists(directory):
        os.makedirs(directory)

    level = logging.INFO
    if os.environ.get('_DEBUG') == '1':
        level = logging.DEBUG
    fmt = '[%(asctime)s %(levelname)s | %(module)s %(funcName)s] %(message)s'
    logging.basicConfig(
        filename=filename, level=logging.DEBUG, format=fmt,
        datefmt="%Y-%M-%d %H:%M:%S")
    console = logging.StreamHandler()
    console.setLevel(level)
    console.setFormatter(logging.Formatter(fmt=fmt, datefmt="%H:%M:%S"))
    logging.getLogger().addHandler(console)


def load_config(filename, env):
    with open(filename, 'r') as f:
        config = json.loads(f.read())

    return config[env]


def get_db(uri, env, database):
    client = pymongo.MongoClient(uri[env])
    db = '%s%s' % (env, database)
    return client[db]


def utc2local(utc_datetime):
    if not utc_datetime:
        return utc_datetime
    # utc_dt = datetime.strptime(utcstr, UTC_FORMAT)
    # local: 东6区
    lo_dt = utc_datetime + timedelta(hours=6)
    return lo_dt.strftime(LOCAL_FORMAT)


def get_deep_start_id(deep):
    if deep in [0, 1, 2, 3]:
        return int(1000 * 10 ** ((deep - 1) * 3))
    elif deep > 3:
        return int(1000 * 10 ** (1+deep))


def convert_time(date_time_str):
    if date_time_str:
        return datetime.strptime(date_time_str, DATA_TIME_FORMAT)
    return None


def convert_time_str(date_time):
    if date_time:
        return date_time.strftime(DATA_TIME_FORMAT)
    return None


def strip_string_name(name):
    if not name:
        return
    if not isinstance(name, str):
        name = str(name)
    name = name.strip()
    split_names = [s.capitalize() if "-" not in s else s for s in name.split(" ")]
    split_names = list(filter(None, split_names))
    name = " ".join(split_names)
    return name


def pull_pickup_station(file_route, from_seller=False):
    wb_obj = load_workbook(file_route, data_only=True)
    ks_sheet = wb_obj.get_sheet_by_name("KS&ward")
    total_count = 0
    for row in list(ks_sheet.iter_rows(min_row=434, max_row=621)):
        total_count += 1
        county, sub_county, area = strip_string_name(
            row[0].value), strip_string_name(row[1].value), strip_string_name(
            row[2].value)
        shop_id = strip_string_name(row[4].value)
        if not shop_id or not shop_id.isdigit():
            continue
        shop_id = int(shop_id)
        # address, shop_id, shop_name = row[3].value, row[4].value, row[5].value
        res_county = bee_common_db.Areas.find_one({"deep": 1, "name": county})
        if not res_county:
            continue
        res_sub_county = bee_common_db.Areas.find_one(
            {"deep": 2, "name": sub_county, "parentId": res_county["_id"]})
        if not res_sub_county:
            continue
        res_area = bee_common_db.Areas.find_one(
            {"deep": 3, "name": area, "parentId": res_sub_county["_id"]})
        if not res_area:
            continue
        station = bee_common_db.PickupStation.find_one_and_update(
            {"sourceStationId": shop_id},
            {"$pull": {"leafAreaIds": res_area["_id"]}})
        if not station:
            continue
        bee_common_db.Areas.update_one(
            {"_id": res_area["_id"]},
            {"$pull": {"pickupIds": shop_id}})
        # if from_seller is True:
        #     member_db.address.update_many(
        #         {"addressType": 2, "pickupStationId": shop_id},
        #         {"$set": {"status": 2}})
        print("update success area=%s,sourceStationId=%s,line=%s" % (
            area, shop_id, total_count))
    print("loop count=%s" % total_count)


def push_pickup_station(file_route):
    wb_obj = load_workbook(file_route, data_only=True)
    ks_sheet = wb_obj.get_sheet_by_name("KS&ward")
    total_count, error_stations = 2, []
    for row in list(ks_sheet.iter_rows(min_row=3, max_row=679)):
        total_count += 1
        county, sub_county, area = strip_string_name(
            row[0].value), strip_string_name(row[1].value), strip_string_name(
            row[2].value)
        shop_id = strip_string_name(row[4].value)
        if not shop_id or not shop_id.isdigit():
            continue
        shop_id = int(shop_id)
        # address, shop_id, shop_name = row[3].value, row[4].value, row[5].value
        res_county = bee_common_db.Areas.find_one({"deep": 1, "name": county})
        if not res_county:
            continue
        res_sub_county = bee_common_db.Areas.find_one(
            {"deep": 2, "name": sub_county, "parentId": res_county["_id"]})
        if not res_sub_county:
            continue
        area_dict = bee_common_db.Areas.find_one(
            {"deep": 3, "name": area, "parentId": res_sub_county["_id"]})
        if area_dict:
            station = bee_common_db.PickupStation.find_one_and_update(
                {"sourceStationId": shop_id},
                {"$addToSet": {"leafAreaIds": area_dict["_id"]}})
            if not station:
                error_stations.append(shop_id)
            bee_common_db.Areas.update_one(
                {"_id": area_dict["_id"]},
                {"$addToSet": {"pickupIds": shop_id}})
        if row[6].value:
            for it in row[6].value.split(","):
                it = strip_string_name(it)
                if it:
                    res_area = bee_common_db.Areas.find_one(
                        {"deep": 3, "name": it, "parentId": res_sub_county["_id"]})
                    if not res_area:
                        continue
                    station = bee_common_db.PickupStation.find_one_and_update(
                        {"sourceStationId": shop_id},
                        {"$addToSet": {"leafAreaIds": res_area["_id"]}})
                    if not station:
                        error_stations.append(shop_id)
                    bee_common_db.Areas.update_one(
                        {"_id": res_area["_id"]},
                        {"$addToSet": {"pickupIds": shop_id}})
                    print(
                        "update pickup area=%s,sourceStationId=%s,line=%s" % (
                            it, shop_id, total_count))
    print("error_stations=%r" % error_stations)


def get_not_match_address(file_route):
    wb_obj = load_workbook(file_route, data_only=True)
    ks_sheet = wb_obj.get_sheet_by_name("KS&ward")
    current_row = 3
    for row in list(ks_sheet.iter_rows(min_row=3, max_row=678)):
        county, sub_county, area = strip_string_name(
            row[0].value), strip_string_name(row[1].value), strip_string_name(
            row[2].value)
        # shop_id = strip_string_name(row[4].value)
        # if not shop_id or not shop_id.isdigit():
        #     continue
        # shop_id = int(shop_id)
        # address, shop_id, shop_name = row[3].value, row[4].value, row[5].value
        res_county = bee_common_db.Areas.find_one({"deep": 1, "name": county})
        if not res_county:
            ks_sheet.cell(current_row, 8).value = county
            current_row += 1
            continue
        res_sub_county = bee_common_db.Areas.find_one(
            {"deep": 2, "name": sub_county, "parentId": res_county["_id"]})
        if not res_sub_county:
            ks_sheet.cell(current_row, 9).value = sub_county
            current_row += 1
            continue
        area_dict = bee_common_db.Areas.find_one(
            {"deep": 3, "name": area, "parentId": res_sub_county["_id"]})
        not_find_area = []
        if not area_dict:
            not_find_area.append(area)
        if row[6].value:
            for it in row[6].value.split(","):
                it = strip_string_name(it)
                if it:
                    res_area = bee_common_db.Areas.find_one(
                        {"deep": 3, "name": it,
                         "parentId": res_sub_county["_id"]})
                    if not res_area:
                        not_find_area.append(it)
        ks_sheet.cell(current_row, 10).value = ",".join(not_find_area)
        current_row += 1
        print(current_row)
    wb_obj.save(file_route)


def get_xed_address(file_name):
    workbook = xlsxwriter.Workbook(file_name)
    worksheet_1 = workbook.add_worksheet("address")
    bold = workbook.add_format(
        {'bold': True, 'align': 'center', 'valign': 'vcenter'})
    other_bold = workbook.add_format({'align': 'center', 'valign': 'vcenter'})
    worksheet_1.write("A1", "ID", bold)
    worksheet_1.write("B1", "County", bold)
    worksheet_1.write("C1", "SubCounty", bold)
    worksheet_1.write("D1", "Word", bold)
    worksheet_1.write("E1", "SupportToDoor", bold)
    worksheet_1.set_column('A:E', 20, other_bold)
    worksheet_1.set_row(0, 20)
    current_row = 2
    for it in bee_common_db.Areas.find({"regionCode": "KE", "deep": 3}).sort(
            [("parentId", 1), ("name", 1)]):
        sub_county = bee_common_db.Areas.find_one(
            {"_id": it["parentId"]})
        county = bee_common_db.Areas.find_one({"_id": sub_county["parentId"]})
        worksheet_1.write('A%d' % current_row, it.get("_id"))
        worksheet_1.write('B%d' % current_row, county["name"])
        worksheet_1.write('C%d' % current_row, sub_county["name"])
        worksheet_1.write('D%d' % current_row, it["name"])
        worksheet_1.write('E%d' % current_row, it.get("isSupportToDoor"))
        current_row += 1
        print(current_row, it.get("_id"))

    workbook.close()


def get_kiliexpress_unmatch_areas():
    areas = []
    for it in bee_common_db.LogisticsAddress.find(
            {"deep": 2, "leafAreaIds": []}).sort([("_id", -1)]):
        city = bee_common_db.LogisticsAddress.find_one({"_id": it["parentId"]})
        areas.append((city["name"], it["name"]))
        print(city["name"], it["name"])
    print(areas)


def update_express_address():
    deep_1_id, deep_2_id = 0, 0
    for it in bee_common_db.LogisticsAddress.find(
            {"deep": 1}).sort([("_id", -1)]).limit(1):
        deep_1_id = it["_id"] + 1
    for it in bee_common_db.LogisticsAddress.find(
            {"deep": 2}).sort([("_id", -1)]).limit(1):
        deep_2_id = it["_id"] + 1
    token = sign_in()["data"]
    for region in get_area_tree(token)["data"]:
        if region.get("country") == "KE":
            for city in region.get("subAreaList"):
                area_2 = bee_common_db.LogisticsAddress.find_one(
                    {"selfAreaId": city["areaId"]})
                if not area_2:
                    print(city, 1)
                    bee_common_db.LogisticsAddress.insert_one(
                        {
                            "_id": deep_1_id,
                            "regionCode": region["code"],
                            "parentId": 8,
                            "selfAreaId": city.get("areaId"),
                            "selfAreaCode": city.get("areaCode"),
                            "providerId": "KiliExpress",
                            "name": city.get("areaNameEn"),
                            "deep": 1,
                            "isSupportToDoor": bool(city.get("supportToDoor")),
                            "isSupportPickup": False,
                            "isLeaf": False,
                            "state": 1,
                            "areaIds": [],
                            "leafAreaIds": [],
                            "pickupIds": []
                        }
                    )
                    deep_1_id += 1
                deep_id_1 = area_2["_id"] if area_2 else deep_1_id
                for area in city.get("subAreaList") or []:
                    if not bee_common_db.LogisticsAddress.find_one(
                            {"selfAreaId": area["areaId"]}):
                        print(area, 2)
                        bee_common_db.LogisticsAddress.insert_one(
                            {
                                "_id": deep_2_id,
                                "regionCode": "KE",
                                "parentId": deep_id_1,
                                "selfAreaId": area["areaId"],
                                "selfAreaCode": area["areaCode"],
                                "providerId": "KiliExpress",
                                "name": area["areaNameEn"],
                                "deep": 2,
                                "isSupportToDoor": bool(area["supportToDoor"]),
                                "isSupportPickup": False,
                                "isLeaf": True,
                                "state": area["status"],
                                "areaIds": [],
                                "leafAreaIds": [],
                                "pickupIds": []
                            }
                        )
                        deep_2_id += 1


if __name__ == "__main__":
    usage = 'python3 Sxx.py prd|dev|test'
    init_logging("release_ready.log")

    if len(sys.argv) < 2:
        logging.error(usage)
        sys.exit(-1)

    env, source = sys.argv[1], sys.argv[2]

    bee_common_db = get_db(K_DB_URL, env, "BeeCommon")

    """
    if source == "xed":
        bee_common_db = get_db(url, env, "BeeCommon")
        from_seller = False
        member_db = {}
    else:
        bee_common_db = get_db(url, env, "Common")
        from_seller = True
        member_db = get_db(url, env, "Member")
    """

    # 更新PickupStation leafAreaIds字段；更新Areas pickupIds字段（减少）
    # pull_pickup_station("./address Update V6.xlsx", from_seller=from_seller)

    # 更新PickupStation leafAreaIds字段；更新Areas pickupIds字段 （增加）
    # push_pickup_station("./address Update V6.xlsx")

    # 获取没有找到自提点对应三级地址
    # get_not_match_address("./address Update V5.1.xlsx")

    # 获取ke站地址excel
    get_xed_address("ke_address.xlsx")

    # 物流端新增地址的处理
    # update_express_address()

    # 获取没有映射上的物流端二级地址
    # get_kiliexpress_unmatch_areas()
