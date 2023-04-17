# -*- coding:utf-8 -*-
import sys
import os
import logging
import pymongo
import json
from datetime import datetime, timedelta
from openpyxl import load_workbook
from treelib import Tree, Node


_DEFAULT_CONFIG_FILE = '../kili_config.json'
_SOUTHX_CONFIG_FILE = '../config.json'
DATA_TIME_FORMAT = "%Y-%m-%dT%H:%M:%S.%fZ"

REGION_MATCH = {1: "Bangladesh", 3: "Myanmar"}
map_pay_method = {
    1: 'Online',
    2: 'COD',
    3: 'Pre'
}

K_DB_URL = {
    "dev": "mongodb://root:KB5NF1T0aP@mongodb-headless.os:27017/admin?replicaSet=rs0",
	"test": "mongodb://root:IdrCgVpHzv@mongo-mongodb-headless.os:27017/admin?replicaSet=rs0&retrywrites=false",
    # "prd": "mongodb://lite-prd",
}


UTC_FORMAT = '%Y-%m-%dT%H:%M:%S.%fZ'
LOCAL_FORMAT = '%Y-%m-%d %H:%M:%S'


def init_logging(filename):
    directory = os.path.dirname(filename)
    if directory != '' and not os.path.exists(directory):
        os.makedirs(directory)

    level = logging.INFO
    if os.environ.get('_DEBUG') == '1':
        level = logging.DEBUG
    fmt = '[%(asctime)s %(levelname)s | %(module)s %(funcName)s] %(message)s'
    logging.basicConfig(
        filename=filename, level=logging.DEBUG, format=fmt,
        datefmt="%Y-%M-%d %H:%M:%S")
    console = logging.StreamHandler()
    console.setLevel(level)
    console.setFormatter(logging.Formatter(fmt=fmt, datefmt="%H:%M:%S"))
    logging.getLogger().addHandler(console)


def load_config(filename, env):
    with open(filename, 'r') as f:
        config = json.loads(f.read())

    return config[env]


def get_db(uri, env, database):
    client = pymongo.MongoClient(uri[env])
    db = '%s%s' % (env, database)
    return client[db]


def utc2local(utc_datetime):
    if not utc_datetime:
        return utc_datetime
    # utc_dt = datetime.strptime(utcstr, UTC_FORMAT)
    # local: 东6区
    lo_dt = utc_datetime + timedelta(hours=6)
    return lo_dt.strftime(LOCAL_FORMAT)


def get_deep_start_id(deep):
    if deep in [0, 1, 2, 3]:
        return int(1000 * 10 ** ((deep - 1) * 3))
    elif deep > 3:
        return int(1000 * 10 ** (1+deep))


def convert_time(date_time_str):
    if date_time_str:
        return datetime.strptime(date_time_str, DATA_TIME_FORMAT)
    return None


def convert_time_str(date_time):
    if date_time:
        return date_time.strftime(DATA_TIME_FORMAT)
    return None


def strip_string_name(name):
    if not name:
        return
    name = name.strip()
    split_names = [s.capitalize() if "-" not in s else s for s in name.split(" ")]
    split_names = list(filter(None, split_names))
    name = " ".join(split_names)
    return name


def add_kili_address(region_code, address_sheet, depth=3, min_row=2):
    # 删除fms，lite，wms Areas地址库数据
    bee_common_db.Areas.delete_many({"regionCode": region_code})
    common_db.Areas.delete_many({"regionCode": region_code})
    wms_common_db.Areas.delete_many({"regionCode": region_code})

    utc_now = datetime.utcnow()
    fms_region = bee_common_db.Region.find_one({"code": region_code})
    lite_region = common_db.region.find_one({"code": region_code})

    deep_0_fms_id, deep_0_lite_id = fms_region["_id"], lite_region["id"]
    data_0 = {
        "areaType": 1,
        "name": fms_region["name"],
        "code": fms_region["callingCode"],
        "postcode": fms_region.get("postcode"),
        "parentId": -1,
        "deep": 0,
        "regionCode": region_code,
        "isSupportToDoor": False,
        "isLeaf": False,
        "state": 1,
        "pickupIds": [],
        "lastUpdatedUserId": 1,
        "lastUpdatedTime": utc_now,
        "supportCod": False
    }
    data_0.update({"_id": deep_0_fms_id, "sort": deep_0_fms_id})
    bee_common_db.Areas.insert_one(data_0)
    wms_common_db.Areas.insert_one(data_0)
    data_0.update({"_id": deep_0_lite_id, "sort": deep_0_lite_id})
    common_db.Areas.insert_one(data_0)

    deep_1_start_id, deep_2_start_id, deep_3_start_id = 0, 0, 0
    deep_4_start_id = 1010000000
    for it in bee_common_db.Areas.find({"deep": 1}).sort([("_id", -1)]).limit(1):
        deep_1_start_id = it["_id"] + 300
    for it in bee_common_db.Areas.find({"deep": 2}).sort([("_id", -1)]).limit(1):
        deep_2_start_id = it["_id"] + 3000
    for it in bee_common_db.Areas.find({"deep": 3}).sort([("_id", -1)]).limit(1):
        deep_3_start_id = it["_id"] + 10000
    for it in bee_common_db.Areas.find({"deep": 4}).sort([("_id", -1)]).limit(1):
        deep_4_start_id = it["_id"] + 100000
    if 0 in [deep_1_start_id, deep_2_start_id, deep_3_start_id]:
        raise Exception("error start id")

    tree = Tree(tree=None, deep=False, node_class=None, identifier=region_code)
    tree.create_node(tag='0', identifier="node-0", data=region_code)
    for row in list(address_sheet.iter_rows(min_row=min_row)):
        area_name_1 = strip_string_name(row[0].value)
        area_name_2 = strip_string_name(row[1].value)
        area_name_3 = strip_string_name(row[2].value)
        area_name_4 = strip_string_name(row[3].value)
        if area_name_1 in [None]:
            continue
        if region_code in ["UG"]:
            area_name_2 = area_name_3
        if region_code == "TZ":
            middleman = area_name_1
            area_name_1 = area_name_3
            area_name_3 = middleman

        if tree.contains(area_name_1) is False:
            deep_1_start_id += 1
            tree.create_node(tag=deep_1_start_id, identifier=area_name_1,
                             parent='node-0', data=area_name_1)
        depth_2_id = area_name_1 + ":" + area_name_2
        if tree.contains(depth_2_id) is False:
            deep_2_start_id += 1
            tree.create_node(tag=deep_2_start_id, identifier=depth_2_id,
                             parent=area_name_1, data=area_name_2)
        if depth < 3:
            continue
        depth_3_id = depth_2_id + ":" + area_name_3
        if tree.contains(depth_3_id) is False:
            deep_3_start_id += 1
            tree.create_node(tag=deep_3_start_id, identifier=depth_3_id,
                             parent=depth_2_id, data=area_name_3)
        if depth < 4:
            continue
        depth_4_id = depth_3_id + ":" + area_name_4
        if tree.contains(depth_4_id) is False:
            deep_4_start_id += 1
            tree.create_node(tag=deep_4_start_id, identifier=depth_4_id,
                             parent=depth_3_id, data=area_name_4)

    for node_1 in tree.children("node-0"):
        data_1 = {"_id": node_1.tag, "areaType": 1, "name": node_1.data,
                  "code": region_code + "1" + str(node_1.tag),
                  "postcode": None, "sort": node_1.tag, "deep": 1,
                  "regionCode": region_code, "isSupportToDoor": False,
                  "isLeaf": False, "state": 1, "pickupIds": [],
                  "lastUpdatedUserId": 1, "lastUpdatedTime": utc_now,
                  "supportCod": False, "parentId": deep_0_fms_id}
        bee_common_db.Areas.insert_one(data_1)
        wms_common_db.Areas.insert_one(data_1)
        data_1["parentId"] = deep_0_lite_id
        common_db.Areas.insert_one(data_1)
        if node_1.is_leaf() is True:
            continue
        for node_2 in tree.children(node_1.identifier):
            data_2 = {
                "_id": node_2.tag,
                "areaType": 1,
                "name": node_2.data,
                "code": region_code + "2" + str(node_2.tag),
                "postcode": None,
                "parentId": node_1.tag,
                "sort": node_2.tag,
                "deep": 2,
                "regionCode": region_code,
                "isSupportToDoor": True,
                "isLeaf": True if depth == 2 else False,
                "state": 1,
                "pickupIds": [],
                "lastUpdatedUserId": 1,
                "lastUpdatedTime": utc_now,
                "supportCod": True
            }
            bee_common_db.Areas.insert_one(data_2)
            common_db.Areas.insert_one(data_2)
            wms_common_db.Areas.insert_one(data_2)
            if node_2.is_leaf() is True:
                continue
            for node_3 in tree.children(node_2.identifier):
                data_3 = {
                    "_id": node_3.tag,
                    "areaType": 1,
                    "name": node_3.data,
                    "code": region_code + "3" + str(node_3.tag),
                    "postcode": None,
                    "parentId": node_2.tag,
                    "sort": node_3.tag,
                    "deep": 3,
                    "regionCode": region_code,
                    "isSupportToDoor": True,
                    "isLeaf": True if depth == 3 else False,
                    "state": 1,
                    "pickupIds": [],
                    "lastUpdatedUserId": 1,
                    "lastUpdatedTime": utc_now,
                    "supportCod": True
                }
                bee_common_db.Areas.insert_one(data_3)
                common_db.Areas.insert_one(data_3)
                wms_common_db.Areas.insert_one(data_3)
                if node_3.is_leaf() is True:
                    continue
                for node_4 in tree.children(node_3.identifier):
                    data_4 = {
                        "_id": node_4.tag,
                        "areaType": 1,
                        "name": node_4.data,
                        "code": region_code + "3" + str(node_4.tag),
                        "postcode": None,
                        "parentId": node_3.tag,
                        "sort": node_4.tag,
                        "deep": 4,
                        "regionCode": region_code,
                        "isSupportToDoor": True,
                        "isLeaf": True if depth == 4 else False,
                        "state": 1,
                        "pickupIds": [],
                        "lastUpdatedUserId": 1,
                        "lastUpdatedTime": utc_now,
                        "supportCod": True
                    }
                    bee_common_db.Areas.insert_one(data_4)
                    common_db.Areas.insert_one(data_4)
                    wms_common_db.Areas.insert_one(data_4)


def add_xed_region_config(region_code, fms_address_level, lite_address_level):
    lite_region = common_db.region.find_one_and_update(
        {"code": region_code}, {"$set": lite_address_level},
        return_document=True)
    if lite_region is None:
        raise Exception("lite this region not support")
    fms_region = bee_common_db.Region.find_one({"code": region_code})
    if not fms_region:
        region_id = 0
        for it in bee_common_db.Region.find({}).sort([("_id", -1)]).limit(1):
            region_id = it["_id"] + 1
        region_data = {
            "_id": region_id,
            "callingCode": lite_region["callingCode"],
            "code": lite_region["code"],
            "name": lite_region["name"],
            "flag": lite_region["flag"],
            "currency": lite_region["currency"],
            "currencyId": lite_region["currencyId"],
            "currencySymbol": lite_region["currencySymbol"],
            "currencyConversion": lite_region["currencyConversion"],
            "language": lite_region["language"],
            "timeZone": lite_region["timeZone"],
            "index": lite_region["index"]
        }
        region_data.update(fms_address_level)
        bee_common_db.Region.insert_one(region_data)
        region_data["id"] = region_id
        del region_data["_id"]
        wms_common_db.Region.insert_one(region_data)
    wms_region = wms_common_db.Region.find_one({"code": region_code})
    if not wms_region:
        fms_region = bee_common_db.Region.find_one({"code": region_code})
        fms_region["id"] = fms_region["_id"]
        del fms_region["_id"]
        wms_common_db.Region.insert_one(fms_region)

    bee_common_db.Region.find_one_and_update(
        {"code": region_code}, {"$set": fms_address_level})
    wms_common_db.Region.find_one_and_update(
        {"code": region_code}, {"$set": fms_address_level})


if __name__ == "__main__":
    usage = 'python3 Sxx.py prd|dev|test region_code'
    init_logging("release_ready.log")
    if len(sys.argv) < 3:
        logging.error(usage)
        sys.exit(-1)

    env = sys.argv[1]
    code = sys.argv[2]
    bee_common_db = get_db(K_DB_URL, env, "BeeCommon")
    bee_logistics_db = get_db(K_DB_URL, env, "BeeLogistics")
    common_db = get_db(K_DB_URL, env, "Common")
    wms_common_db = get_db(K_DB_URL, env, "WmsCommon")
    depth, from_row, address_sheet = 3, 2, None
    fms_address_level = {
        "addressLevel": 3, "addressNameMap": {
            "1": "Province", "2": "City", "3": "Area"},
        "LogisticsProviders": []}
    lite_address_level = {
        "addressLevel": 3, "addressNameMap": {
            "1": "Province", "2": "City", "3": "Area"}}
    if code in ["UG", "TZ"]:
        wb_obj = load_workbook("region_address_excel/UG&TZ-地址库.xlsx")
        if code == "UG":
            address_sheet = wb_obj.get_sheet_by_name("乌干达")
            fms_address_level = {
                "addressLevel": 2, "addressNameMap": {
                    "1": "Province", "2": "District"},
                "LogisticsProviders": []}
            lite_address_level = {
                "addressLevel": 2, "addressNameMap": {
                    "1": "Province", "2": "District"}}
            depth = 2
        else:
            address_sheet = wb_obj.get_sheet_by_name("坦桑")
    elif code == "NG":
        wb_obj = load_workbook("region_address_excel/NG-地址库.xlsx")
        address_sheet = wb_obj.get_sheet_by_name("地址库")
        fms_address_level = {
            "addressLevel": 3, "addressNameMap": {
                "1": "Region", "2": "State", "3": "LGA", "4": "Ward/City"},
            "LogisticsProviders": []}
        lite_address_level = {
            "addressLevel": 3, "addressNameMap": {
                "1": "Region", "2": "State", "3": "LGA", "4": "Ward/City"}}
        depth = 4
        from_row = 3

    elif code == "EG":
        wb_obj = load_workbook("region_address_excel/EG-地址库（埃及）.xlsx")
        address_sheet = wb_obj.get_sheet_by_name("Site coverage list")
        fms_address_level = {
            "addressLevel": 3, "addressNameMap": {
                "1": "County", "2": "SubCounty", "3": "Ward"},
            "LogisticsProviders": []}
        lite_address_level = {
             "addressLevel": 3, "addressNameMap": {
                "1": "County", "2": "SubCounty", "3": "Ward"}}
    elif code == "MA":
        wb_obj = load_workbook("region_address_excel/MA-地址库（摩洛哥）.xlsx")
        address_sheet = wb_obj.get_sheet_by_name("地址库")
    add_xed_region_config(code, fms_address_level, lite_address_level)
    add_kili_address(code, address_sheet, depth=depth, min_row=from_row)
    print("add_kili_address success")

