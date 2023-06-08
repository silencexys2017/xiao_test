import json
import logging
import sys
import os
import pymongo
from openpyxl import load_workbook

K_DB_URL = {
    "dev": "mongodb://root:KB5NF1T0aP@mongodb-headless.os:27017/admin?replicaSet=rs0",
	"test": "mongodb://root:IdrCgVpHzv@mongo-mongodb-headless.os:27017/admin?replicaSet=rs0&retrywrites=false",
    "prd": ""
}


def init_logging(filename):
    directory = os.path.dirname(filename)
    if directory != '' and not os.path.exists(directory):
        os.makedirs(directory)

    level = logging.INFO
    if os.environ.get('_DEBUG') == '1':
        level = logging.DEBUG
    fmt = '[%(asctime)s %(levelname)s | %(module)s %(funcName)s] %(message)s'
    logging.basicConfig(
        filename=filename, level=logging.DEBUG, format=fmt,
        datefmt="%Y-%M-%d %H:%M:%S")
    console = logging.StreamHandler()
    console.setLevel(level)
    console.setFormatter(logging.Formatter(fmt=fmt, datefmt="%H:%M:%S"))
    logging.getLogger().addHandler(console)


def get_db(uri, env, database):
    client = pymongo.MongoClient(uri[env])
    db = '%s%s' % (env, database)
    return client[db]


def strip_string_name(name):
    if name is None:
        return
    if not isinstance(name, str):
        name = str(name)
    name = name.strip()
    return name


def add_pickup_station(file_route):
    wb_obj = load_workbook(file_route, data_only=True)
    ks_sheet = wb_obj.get_sheet_by_name("Sheet1")
    start_id = None
    for it in bee_common_db.PickupStation.find().sort([("_id", -1)]).limit(1):
        start_id = it["_id"] + 1
    pickup_station_ids = set()
    for row in list(ks_sheet.iter_rows(min_row=2, max_row=1515)):
        station_id = int(strip_string_name(row[0].value))
        if station_id in pickup_station_ids:
            print(station_id)
        pickup_station_ids.add(station_id)
        station = bee_common_db.PickupStation.find_one(
            {"sourceStationId": station_id})
        lite_station = common_db.PickupStation.find_one(
            {"sourceStationId": station_id})
        if station:
            if not lite_station:
                common_db.PickupStation.insert_one(station)
            continue
        print(station_id)
        latitude = strip_string_name(row[10].value)
        if "N" in latitude:
            latitude = None
        else:
            latitude = float(latitude)
        longitude = strip_string_name(row[11].value)
        if "N" in longitude:
            longitude = None
        else:
            longitude = float(longitude)
        is_support_big = strip_string_name(row[4].value)
        if "N" in is_support_big:
            is_support_big = None
        business_type = strip_string_name(row[7].value)
        if "N" in business_type:
            business_type = 3
        else:
            business_type = int(business_type)

        data = {
            "_id": start_id,
            "sourceStationId": station_id,
            "status": int(strip_string_name(row[1].value)),
            "name": strip_string_name(row[2].value),
            "address": strip_string_name(row[3].value),
            "isSupportBigPackage": is_support_big,
            "areaId": int(strip_string_name(row[5].value)),
            "areaCode": strip_string_name(row[6].value),
            "businessType": business_type,
            "pickingCode": strip_string_name(row[8].value),
            "businessTime": strip_string_name(row[9].value),
            "latitude": latitude,
            "longitude": longitude,
            "regionCode": strip_string_name(row[12].value),
            "county": strip_string_name(row[14].value),
            "town": strip_string_name(row[13].value),
            "phone": strip_string_name(row[15].value),
            "leafAreaIds": []
        }
        print(start_id, json.dumps(data))
        bee_common_db.PickupStation.insert_one(data)
        common_db.PickupStation.insert_one(data)
        start_id += 1
    print(len(pickup_station_ids))


if __name__ == "__main__":
    usage = 'python3 Sxx.py prd|dev|test'
    init_logging("release_ready.log")

    if len(sys.argv) < 2:
        logging.error(usage)
        sys.exit(-1)

    env = sys.argv[1]
    bee_common_db = get_db(K_DB_URL, env, "BeeCommon")
    common_db = get_db(K_DB_URL, env, "Common")
    add_pickup_station("./自提点.xlsx")