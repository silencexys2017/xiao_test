# -*- coding:utf-8 -*-
import sys
import os
import logging
import pymongo
import json
from openpyxl import load_workbook
from datetime import timedelta

_DEFAULT_CONFIG_FILE = '../config.json'

REGION_MATCH = {1: "Bangladesh", 3: "Myanmar"}
MONTH = ["January", "February", "March", "April", "May", "June", "July",
         "August", "September", "October", "November", "December"]
map_pay_method = {
    1: 'Online',
    2: 'COD',
    3: 'Pre'
}
"""
序号，XE订单号，商户订单号，SKU ID，商品数量， 商品的中创建时间，国内快递单号，目的国，
最后更新时间，状态，是否打印面单, 商品的中文名"""

map_state = {
    1: '待收件',
    2: '已收件',
    3: '已提运',
    4: '已发出',
    5: '待接件',
    6: '待派送',
    7: '尾程派送中',
    8: '已交付',
    -1: '已取消',
    -2: '已拒收',
    -3: '拒揽件'
}

UTC_FORMAT = '%Y-%m-%dT%H:%M:%S.%fZ'
LOCAL_FORMAT = '%Y-%m-%d %H:%M:%S'


def init_logging(filename):
    directory = os.path.dirname(filename)
    if directory != '' and not os.path.exists(directory):
        os.makedirs(directory)

    level = logging.INFO
    if os.environ.get('_DEBUG') == '1':
        level = logging.DEBUG
    fmt = '[%(asctime)s %(levelname)s | %(module)s %(funcName)s] %(message)s'
    logging.basicConfig(
        filename=filename, level=logging.DEBUG, format=fmt,
        datefmt="%Y-%M-%d %H:%M:%S")
    console = logging.StreamHandler()
    console.setLevel(level)
    console.setFormatter(logging.Formatter(fmt=fmt, datefmt="%H:%M:%S"))
    logging.getLogger().addHandler(console)


def load_config(filename, env):
    with open(filename, 'r') as f:
        config = json.loads(f.read())

    return config[env]


def get_db(config, env, database):
    uri = config['mongodb']['uri']
    client = pymongo.MongoClient(uri)
    db = '%s%s' % (env, database)
    return client[db]


def utc2local(utc_datetime):
    if not utc_datetime:
        return utc_datetime
    # utc_dt = datetime.strptime(utcstr, UTC_FORMAT)
    # local: 东6区
    lo_dt = utc_datetime + timedelta(hours=6)
    return lo_dt.strftime(LOCAL_FORMAT)


def update_excel_data(wb_obj):
    # 使用指定工作表
    sheet = wb_obj.active  # 当前激活的工作表
    # sheet = excel.get_sheet_by_name('Sheet1')
    # 读取所有数据
    # print(list(sheet.values))  # sheet.values 生成器
    # print(sheet.max_column)  # 最大列数
    # print(sheet.max_row)  # 最大行数

    # 按行读取
    # for row in sheet.iter_rows(min_row=1, min_col=1, max_col=3, max_row=3):
    #     print(row)

    # 读取标题行
    # for row in sheet.iter_rows(max_row=1):
    #     title_row = [cell.value for cell in row]
    #     print(title_row)
    # 读取标题行以外数据
    # for row in sheet.iter_rows(min_row=2):
    #     row_data = [cell.value for cell in row]
    #     print(row_data)
    #     for cell in row:
    #         print(cell.value, cell.column, cell.row)

    for row in list(sheet.iter_rows(min_row=2)):
        order_no = row[0].value
        order = logistics_db.LogisticsOrder.find_one({"orderNo": order_no})
        if not order:
            continue
        pg = order.get("supplierList") or []
        sheet.cell(row[0].row, 28).value = utc2local(pg[0].get("receivedAt"))
        sheet.cell(row[0].row, 29).value = utc2local(
            order.get("createdAt"))
        sheet.cell(row[0].row, 30).value = utc2local(
            order.get("bePickupAt"))
        sheet.cell(row[0].row, 31).value = utc2local(
            order.get("localReceivedAt"))
        td = order.get("terminalDelivery") or {}
        sheet.cell(row[0].row, 32).value = utc2local(
            td.get("startDeliveryTime"))
        sheet.cell(row[0].row, 33).value = utc2local(
            order.get("closedAt"))

    # 读取单元格数据
    # print(sheet['A1'].value)
    # print(sheet.cell(1, 1).value)  # 索引从1开始

    # 写入单元格
    # sheet['F2'] = 'PASS'
    # 整行写入
    # new_row = ['post-xml接口', 'post', 'https://httpbin.org/post']
    # sheet.append(new_row)


if __name__ == "__main__":
    usage = 'python3 Sxx.py prd|dev|test'

    init_logging("xed-order-export.log")
    if len(sys.argv) < 2:
        logging.error(usage)
        sys.exit(-1)

    env = sys.argv[1]
    config = load_config(_DEFAULT_CONFIG_FILE, env)
    logistics_db = get_db(config, env, "BeeLogistics")
    FILE_NAME = "订单汇总0308.xlsx"

    wb_obj = load_workbook(FILE_NAME)

    update_excel_data(wb_obj)
    # 保存文件，也可覆盖原文件
    wb_obj.save(FILE_NAME)


