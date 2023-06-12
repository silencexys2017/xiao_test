# -*- coding:utf-8 -*-
import sys
import os
import logging
import pymongo
import json
from datetime import datetime, timedelta
from openpyxl import load_workbook
from prd_kiliexpress import get_pick_up_stations, get_area_tree, sign_in


_DEFAULT_CONFIG_FILE = '../kili_config.json'
_SOUTHX_CONFIG_FILE = '../config.json'
DATA_TIME_FORMAT = "%Y-%m-%dT%H:%M:%S.%fZ"

REGION_MATCH = {1: "Bangladesh", 3: "Myanmar"}
map_pay_method = {
    1: 'Online',
    2: 'COD',
    3: 'Pre'
}

X_DB_URL = {
    "dev": "mongodb://pf_dev_dbo:Bp2Q5j3Jb2cmQvn8L4kW@mongodb-paas-service/admin?replicaSet=rs0",
	"test": "mongodb://pf_test_dbo:3f4k8aDHeQJBKmd3z7c9@159.138.90.30:30734/admin",
	# "prd": "mongodb://pf_prd_dbo:m2w9ZNZP4gUsx9b2hu6L@159.138.90.30:30734/admin",
    "prd": "mongodb://pf_prd_dbo:m2w9ZNZP4gUsx9b2hu6L@mongodb-paas-service/admin?replicaSet=rs0"
}

K_DB_URL = {
    "dev": "mongodb://root:ty8NaMSWoM@10.0.1.103:32346/admin?readPreference=primaryPreferred",
	"test": "mongodb://root:ty8NaMSWoM@10.0.1.103:32346/admin",
	"prd": "mongodb://sk2:Qmvz84mtswuMJz8Kk90zBM7UbdW@ebes-db.cluster-cvxmlpsy4xpw.eu-central-1.docdb.amazonaws.com:27017/admin?authSource=admin"
}


UTC_FORMAT = '%Y-%m-%dT%H:%M:%S.%fZ'
LOCAL_FORMAT = '%Y-%m-%d %H:%M:%S'


def init_logging(filename):
    directory = os.path.dirname(filename)
    if directory != '' and not os.path.exists(directory):
        os.makedirs(directory)

    level = logging.INFO
    if os.environ.get('_DEBUG') == '1':
        level = logging.DEBUG
    fmt = '[%(asctime)s %(levelname)s | %(module)s %(funcName)s] %(message)s'
    logging.basicConfig(
        filename=filename, level=logging.DEBUG, format=fmt,
        datefmt="%Y-%M-%d %H:%M:%S")
    console = logging.StreamHandler()
    console.setLevel(level)
    console.setFormatter(logging.Formatter(fmt=fmt, datefmt="%H:%M:%S"))
    logging.getLogger().addHandler(console)


def load_config(filename, env):
    with open(filename, 'r') as f:
        config = json.loads(f.read())

    return config[env]


def get_db(uri, env, database):
    client = pymongo.MongoClient(uri[env])
    db = '%s%s' % (env, database)
    return client[db]


def utc2local(utc_datetime):
    if not utc_datetime:
        return utc_datetime
    # utc_dt = datetime.strptime(utcstr, UTC_FORMAT)
    # local: 东6区
    lo_dt = utc_datetime + timedelta(hours=6)
    return lo_dt.strftime(LOCAL_FORMAT)


def get_deep_start_id(deep):
    if deep in [0, 1, 2, 3]:
        return int(1000 * 10 ** ((deep - 1) * 3))
    elif deep > 3:
        return int(1000 * 10 ** (1+deep))


def convert_time(date_time_str):
    if date_time_str:
        return datetime.strptime(date_time_str, DATA_TIME_FORMAT)
    return None


def convert_time_str(date_time):
    if date_time:
        return date_time.strftime(DATA_TIME_FORMAT)
    return None

def strip_string_name(name):
    if not name:
        return
    name = name.strip()
    split_names = [s.capitalize() if "-" not in s else s for s in name.split(" ")]
    split_names = list(filter(None, split_names))
    name = " ".join(split_names)
    return name


def import_ke_address_from_excel(file_route):
    wb_obj = load_workbook(file_route, data_only=True)
    xed_sheet = wb_obj.get_sheet_by_name('new XED')
    xed_areas = {}
    error_kl_data = {"county": [], "town": []}
    for row in list(xed_sheet.iter_rows(min_row=2)):
        county, sub_county, area = strip_string_name(
            row[0].value), strip_string_name(row[1].value), strip_string_name(
            row[2].value)
        k_county, k_town = strip_string_name(row[4].value), strip_string_name(
            row[5].value)
        # if not bee_common_db.KiliCity.find_one({"name": k_county}):
        #     xed_sheet.cell(row[0].row, 10).value = k_county
        #     error_kl_data["county"].append(k_county)
        # if not bee_common_db.KiliArea.find_one({"name": k_town}):
        #     xed_sheet.cell(row[0].row, 11).value = k_town
        #     error_kl_data["town"].append(k_town)
        res_struct = {"kiLiAddress": [k_county, k_town], "shopIds": []}
        if xed_areas.get(county):
            if xed_areas[county].get(sub_county):
                xed_areas[county][sub_county][area] = res_struct
            else:
                xed_areas[county][sub_county] = {area: res_struct}
        else:
            xed_areas[county] = {sub_county: {area: res_struct}}
    xed_areas["errorKiliName"] = error_kl_data

    mall_areas, map_address = {}, {}
    not_match_town, not_match_ward = [], []
    mall_sheet = wb_obj.get_sheet_by_name("Mall adress")
    mall_sub_county = {}
    for row in list(mall_sheet.iter_rows(min_row=2)):
        county, sub_county, area = strip_string_name(
            row[0].value), strip_string_name(row[1].value), strip_string_name(
            row[2].value)
        k_county, k_town = strip_string_name(row[4].value), strip_string_name(
            row[5].value)
        map_address[area] = (county, sub_county)
        # if not bee_common_db.KiliCityApi.find_one({"nameEn": k_county}):
        #     error_kl_data["county"].append(k_county)
        #     mall_sheet.cell(row[0].row, 8).value = k_county
        # if not bee_common_db.KiliAreaApi.find_one({"nameEn": k_town}):
        #     mall_sheet.cell(row[0].row, 9).value = k_town
        #     error_kl_data["town"].append(k_town)
        if len(strip_string_name(row[10].value)) > 7:
            sup_door = False
        else:
            sup_door = True

        res_struct = {
            "kiLiAddress": [k_county, k_town], "shopIds": [], "type": 1,
            "supportDoor": sup_door}
        if mall_areas.get(county):
            if mall_areas[county].get(sub_county):
                mall_areas[county][sub_county][area] = res_struct
            else:
                mall_areas[county][sub_county] = {area: res_struct}
        else:
            mall_areas[county] = {sub_county: {area: res_struct}}

        if not xed_areas[county].get(sub_county):
            mall_sub_county[county+";"+sub_county] = (county, sub_county)
            for vale in mall_areas[county][sub_county].values():
                vale["type"] = 2
                # mall_sheet.cell(row[0].row, 10).value = "Y"
            not_match_town.append((county, sub_county, area))
        elif xed_areas[county][sub_county].get(area) is None:
            mall_areas[county][sub_county][area]["type"] = 2
            # mall_sheet.cell(row[0].row, 10).value = "Y"
            not_match_ward.append((county, sub_county, area))
    error_kl_data["county"] = list(set(error_kl_data["county"]))
    error_kl_data["town"] = list(set(error_kl_data["town"]))
    print(error_kl_data)
    mall_areas["notMatchTown"] = not_match_town
    mall_areas["notMatchWard"] = not_match_ward
    mall_areas["mallSubCounty"] = mall_sub_county
    print(len(not_match_town), len(not_match_ward))

    ks_sheet = wb_obj.get_sheet_by_name("KS&ward")
    shop_not_match = []
    map_wards = []
    for row in list(ks_sheet.iter_rows(min_row=3)):
        county, sub_county, area = strip_string_name(
            row[0].value), strip_string_name(row[1].value), strip_string_name(
            row[2].value)
        shop_id = strip_string_name(row[4].value)
        shop_id = int(shop_id) if shop_id else 0
        # address, shop_id, shop_name = row[3].value, row[4].value, row[5].value
        if row[6].value:
            for it in row[6].value.split(","):
                it = strip_string_name(it)
                if it:
                    map_wards.append(it+"||"+str(shop_id))
        try:
            mall_areas[county][sub_county][area]["shopIds"].append(shop_id)
        except Exception:
            try:
                xed_areas[county][sub_county][area]["shopIds"].append(shop_id)
                print(shop_id)
            except Exception as exp:
                pass
            # ks_sheet.cell(row[0].row, 8).value = "Y"
            shop_not_match.append((county, sub_county, area))
    start_row = 434
    not_map_ward = []
    for item in map_wards:
        ward, shop_id = item.split("||")
        if map_address.get(ward):
            ks_sheet.cell(start_row, 3).value = ward
            ks_sheet.cell(start_row, 1).value = map_address[ward][0]
            ks_sheet.cell(start_row, 2).value = map_address[ward][1]
            ks_sheet.cell(start_row, 5).value = shop_id
            start_row += 1
        else:
            not_map_ward.append(item)
    for item in not_map_ward:
        ward, shop_id = item.split("||")
        ks_sheet.cell(start_row, 3).value = ward
        ks_sheet.cell(start_row, 5).value = shop_id
        start_row += 1

    xed_areas["shopNotMatch"] = shop_not_match

    with open("xed_address.json", 'w', encoding="utf-8") as write_f:
        write_f.write(json.dumps(xed_areas, indent=4, ensure_ascii=False))
    with open("mall_address.json", 'w', encoding="utf-8") as write_f:
        write_f.write(json.dumps(mall_areas, indent=4, ensure_ascii=False))
    wb_obj.save(file_route)


def add_kili_ke_address():
    bee_common_db.Areas.delete_many({})
    utc_now = datetime.utcnow()
    deep_0_start_id = get_deep_start_id(0)
    deep_1_start_id = get_deep_start_id(1)
    deep_2_start_id = get_deep_start_id(2)
    deep_3_start_id = get_deep_start_id(3)
    cycles = 1
    for region in bee_common_db.Region.find({}):
        region_code = region.get("code")
        if region_code not in ["BD", "CN", "MM", "KE"]:
            continue
        region_id = region["_id"] if region_code == "KE" else region["_id"]
        data_0 = {
            "_id": region_id,
            "areaType": 1,
            "name": region["name"],
            "code": region["callingCode"],
            "postcode": region.get("postcode"),
            "parentId": -1,
            "sort": region_id,
            "deep": 0,
            "regionCode": region_code,
            "isSupportToDoor": False,
            "isLeaf": False,
            "state": 1,
            "pickupIds": [],
            "lastUpdatedUserId": 1,
            "lastUpdatedTime": utc_now,
            "supportCod": False
        }
        bee_common_db.Areas.insert_one(data_0)
        if cycles > 1:
            deep_1_start_id = deep_1_start_id + 100
            deep_2_start_id = deep_2_start_id + 1000
            deep_3_start_id = deep_3_start_id + 10000
        cycles += 1
        if region_code == "KE":
            with open("mall_address.json", 'r') as f:
                import_ke_address(
                    json.load(f), region_id, region_code, deep_1_start_id,
                    deep_2_start_id, deep_3_start_id, utc_now)

            continue
        for state in bee_common_db.State.find(
                {"regionId": region_id}).sort([("name", 1)]):
            deep_1_start_id += 1
            code = region_code + "1" + str(deep_1_start_id)
            data_1 = {
                "_id": deep_1_start_id,
                "areaType": 1,
                "name": state["name"],
                "code": code,
                "postcode": state.get("postcode"),
                "parentId": region_id,
                "sort": deep_1_start_id,
                "deep": 1,
                "regionCode": region_code,
                "isSupportToDoor": False,
                "isLeaf": False,
                "state": 1,
                "pickupIds": [],
                "lastUpdatedUserId": 1,
                "lastUpdatedTime": utc_now,
                "originalId": state["_id"],
                "supportCod": False
            }
            bee_common_db.Areas.insert_one(data_1)
            for city in bee_common_db.City.find(
                    {"stateId": state["_id"]}).sort([("name", 1)]):
                deep_2_start_id += 1
                code = region_code + "2" + str(deep_2_start_id)
                data_2 = {
                    "_id": deep_2_start_id,
                    "areaType": 1,
                    "name": city["name"],
                    "code": code,
                    "postcode": city.get("postcode"),
                    "parentId": deep_1_start_id,
                    "sort": deep_2_start_id,
                    "deep": 2,
                    "regionCode": region_code,
                    "isSupportToDoor": False,
                    "isLeaf": False,
                    "state": 1,
                    "pickupIds": [],
                    "lastUpdatedUserId": 1,
                    "lastUpdatedTime": utc_now,
                    "originalId": city["_id"],
                    "supportCod": False
                }
                bee_common_db.Areas.insert_one(data_2)
                bee_logistics_db.ExfcsAddressMapping.update_one(
                    {"pfCityId": city["_id"]},
                    {"$set": {"pfCityId": deep_2_start_id}})
                for area in bee_common_db.Area.find(
                        {"cityId": city["_id"]}).sort([("name", 1)]):
                    deep_3_start_id += 1
                    code = region_code + "3" + str(deep_3_start_id)
                    data_3 = {
                        "_id": deep_3_start_id,
                        "areaType": 1,
                        "name": area["name"],
                        "code": code,
                        "postcode": area.get("postcode"),
                        "parentId": deep_2_start_id,
                        "sort": deep_3_start_id,
                        "deep": 3,
                        "regionCode": region_code,
                        "isSupportToDoor": False,
                        "isLeaf": True,
                        "state": 1,
                        "pickupIds": [],
                        "lastUpdatedUserId": 1,
                        "lastUpdatedTime": utc_now,
                        "originalId": area["_id"],
                        "supportCod": False
                    }
                    bee_common_db.Areas.insert_one(data_3)
                    bee_logistics_db.ECourierAddressMapping.update_one(
                        {"pfAreaId": area["_id"]},
                        {"$addToSet": {"pfAreaId": deep_3_start_id}})
                    bee_logistics_db.EDeshAddressMapping.update_one(
                        {"areaId": area["_id"]},
                        {"$set": {"areaId": deep_3_start_id}})


def add_xed_region_config():
    for it in bee_common_db.Region.find({}):
        query = {"code": it["code"]}
        address_name = {
            "addressLevel": 3,
            "addressNameMap": {"1": "State", "2": "City", "3": "Area"},
            "LogisticsProviders": []}
        if it["code"] == "KE":
            address_name = {
                "addressLevel": 3,
                "addressNameMap": {
                "1": "County", "2": "Subcounty/Town", "3": "Area(ward)"},
                "LogisticsProviders": ["KiliExpress"]}
        elif it["code"] == "BD":
            address_name = {
                "addressLevel": 3,
                "addressNameMap": {"1": "State", "2": "City", "3": "Area"},
                "LogisticsProviders": ["eCourier", "E-Desh"]}
        elif it["code"] == "CN":
            address_name = {
                "addressLevel": 2,
                "addressNameMap": {"1": "State", "2": "City"},
                "LogisticsProviders": []}
        bee_common_db.Region.update_one(query, {"$set": address_name})


def add_kili_pickup_station():
    bee_common_db.PickupStation.delete_many({})
    start_id, shop_area_map = 1, {}
    with open("shop_area_map.json", 'r') as f:
        shop_area_map = json.load(f)
    token = sign_in()["data"]
    for it in bee_common_db.KiliAreaApi.find({}):
        county = bee_common_db.KiliCityApi.find_one(
            {"id": it["parentId"]})["nameEn"]
        for station in get_pick_up_stations(
                it["code"], token=token).get("data") or []:
            data = {
                "_id": start_id,
                "sourceStationId": station["id"],
                "status": 1,
                "name": station["stationName"],
                "address": station["address"],
                "isSupportBigPackage": station.get("isSupportBigPackage"),
                "areaId": station["areaId"],
                "areaCode": station["areaCode"],
                "businessType": station["businessType"],
                "pickingCode": station["pickingCode"],
                "businessTime": station["businessTime"],
                "latitude": station["latitude"],
                "longitude": station["longitude"],
                "regionCode": "KE",
                "county": county,
                "town": it["nameEn"],
                "phone": station.get("phoneNumber"),
                "leafAreaIds": shop_area_map.get(str(station["id"]), [])
            }
            bee_common_db.PickupStation.insert_one(data)
            ki_li_data["stations"].append(data)
            start_id += 1


def import_ke_address(
        ke_json, region_id, region_code, deep_1_start_id, deep_2_start_id,
        deep_3_start_id, utc_now):
    mall_sub_county = ke_json["mallSubCounty"]
    shop_area_map = {}
    for state, s_item in ke_json.items():
        if state in ["notMatchTown", "notMatchWard", "mallSubCounty"]:
            continue
        deep_1_start_id += 1
        code = region_code + "1" + str(deep_1_start_id)
        data_1 = {
            "_id": deep_1_start_id,
            "areaType": 1,
            "name": state,
            "code": code,
            "postcode": None,
            "parentId": region_id,
            "sort": deep_1_start_id,
            "deep": 1,
            "regionCode": region_code,
            "isSupportToDoor": False,
            "isLeaf": False,
            "state": 1,
            "pickupIds": [],
            "lastUpdatedUserId": 1,
            "lastUpdatedTime": utc_now,
            "supportCod": False
        }
        bee_common_db.Areas.insert_one(data_1)

        for city, c_item in s_item.items():
            deep_2_start_id += 1
            code = region_code + "2" + str(deep_2_start_id)
            if mall_sub_county.get(state+";"+city):
                area_type = 2
            else:
                area_type = 1
            data_2 = {
                "_id": deep_2_start_id,
                "areaType": area_type,
                "name": city,
                "code": code,
                "postcode": None,
                "parentId": deep_1_start_id,
                "sort": deep_2_start_id,
                "deep": 2,
                "regionCode": region_code,
                "isSupportToDoor": False,
                "isLeaf": False,
                "state": 1,
                "pickupIds": [],
                "lastUpdatedUserId": 1,
                "lastUpdatedTime": utc_now,
                "supportCod": False
            }
            bee_common_db.Areas.insert_one(data_2)
            for area, a_item in c_item.items():
                deep_3_start_id += 1
                code = region_code + "3" + str(deep_3_start_id)
                data_3 = {
                    "_id": deep_3_start_id,
                    "areaType": a_item["type"],
                    "name": area,
                    "code": code,
                    "postcode": None,
                    "parentId": deep_2_start_id,
                    "sort": deep_3_start_id,
                    "deep": 3,
                    "regionCode": region_code,
                    "isSupportToDoor": a_item["supportDoor"],
                    "isLeaf": True,
                    "state": 1,
                    "pickupIds": a_item["shopIds"],
                    "lastUpdatedUserId": 1,
                    "lastUpdatedTime": utc_now,
                    "supportCod": False
                }
                for s_id in a_item["shopIds"]:
                    if shop_area_map.get(s_id):
                        shop_area_map[s_id].append(deep_3_start_id)
                    else:
                        shop_area_map[s_id] = [deep_3_start_id]
                bee_common_db.Areas.insert_one(data_3)
                kk_ad = a_item["kiLiAddress"]
                ct = bee_common_db.LogisticsAddress.find_one_and_update(
                    {"name": kk_ad[0], "deep": 1},
                    {"$addToSet": {"areaIds": deep_1_start_id}})
                if ct:
                    bee_common_db.LogisticsAddress.update_one(
                        {"parentId": ct["_id"], "name": kk_ad[1]},
                        {"$addToSet": {
                            "leafAreaIds": deep_3_start_id,
                            "areaIds": {
                                "$each": [deep_2_start_id, deep_3_start_id]}}}
                    )
    with open("shop_area_map.json", 'w', encoding="utf-8") as write_f:
        write_f.write(json.dumps(shop_area_map, indent=4, ensure_ascii=False))


def add_kili_express_address():
    bee_common_db.LogisticsAddress.delete_many({})
    region = bee_common_db.Region.find_one({"code": "KE"})
    deep_1_id = get_deep_start_id(1)
    deep_2_id = get_deep_start_id(2)
    for city in bee_common_db.KiliCityApi.find().sort([("name", 1)]):
        deep_1_id += 1
        bee_common_db.LogisticsAddress.insert_one(
            {
                "_id": deep_1_id,
                "regionCode": region["code"],
                "parentId": 8,
                "selfAreaId": city["id"],
                "selfAreaCode": city["code"],
                "providerId": "KiliExpress",
                "name": city["nameEn"],
                "deep": 1,
                "isSupportToDoor": bool(city["supportToDoor"]),
                "isSupportPickup": False,
                "isLeaf": False,
                "state": 1,
                "areaIds": [],
                "leafAreaIds": [],
                "pickupIds": []
            }
        )
        for area in bee_common_db.KiliAreaApi.find(
                {"parentId": city["id"]}).sort([("name", 1)]):
            deep_2_id += 1
            bee_common_db.LogisticsAddress.insert_one(
                {
                    "_id": deep_2_id,
                    "regionCode": "KE",
                    "parentId": deep_1_id,
                    "selfAreaId": area["id"],
                    "selfAreaCode": area["code"],
                    "providerId": "KiliExpress",
                    "name": area["nameEn"],
                    "deep": 2,
                    "isSupportToDoor": bool(city["supportToDoor"]),
                    "isSupportPickup": False,
                    "isLeaf": True,
                    "state": area["status"],
                    "areaIds": [],
                    "leafAreaIds": [],
                    "pickupIds": []
                }
            )


def create_kili_express_address():
    bee_common_db.KiliCityApi.delete_many({})
    bee_common_db.KiliAreaApi.delete_many({})
    token = sign_in()["data"]
    for region in get_area_tree(token)["data"]:
        if region.get("country") == "KE":
            for city in region.get("subAreaList"):
                bee_common_db.KiliCityApi.insert_one({
                    "id": city.get("areaId"),
                    "parentId": city.get("parentId"),
                    "name": city.get("areaName"),
                    "nameEn": city.get("areaNameEn"),
                    "country": city.get("country"),
                    "code": city.get("areaCode"),
                    "sort": city.get("areaSort"),
                    "status": city.get("status"),
                    "supportToDoor": city.get("supportToDoor")
                })

                for area in city.get("subAreaList") or []:
                    bee_common_db.KiliAreaApi.insert_one({
                        "id": area.get("areaId"),
                        "parentId": area.get("parentId"),
                        "name": area.get("areaName"),
                        "nameEn": area.get("areaNameEn"),
                        "country": area.get("country"),
                        "code": area.get("areaCode"),
                        "sort": area.get("areaSort"),
                        "status": area.get("status"),
                        "supportToDoor": area.get("supportToDoor")
                    })


if __name__ == "__main__":
    usage = 'python3 Sxx.py prd|dev|test'
    init_logging("release_ready.log")
    if len(sys.argv) < 2:
        logging.error(usage)
        sys.exit(-1)

    env = sys.argv[1]
    # config = load_config(_DEFAULT_CONFIG_FILE, env)
    # southx_config = load_config(_SOUTHX_CONFIG_FILE, env)
    bee_common_db = get_db(X_DB_URL, env, "BeeCommon")
    bee_logistics_db = get_db(X_DB_URL, env, "BeeLogistics")

    # 创建kili物流api地址库
    # create_kili_express_address()
    # print("create_kili_express_address success")

    # 创建kili物流地址库
    # add_kili_express_address()
    # print("add_kili_express_address success")

    # 将excel地址库数据，存到json （xed_address.json，mall_address.json）
    # import_ke_address_from_excel("./address Update V5.xlsx")

    # 地址库导入到数据库
    # add_kili_ke_address()
    # print("add_kili_ke_address success")

    # 增加国家配置
    # add_xed_region_config()
    # print("add_xed_region_config success")
    ki_li_data = {"stations": [], "areas": []}

    # 自提点写入数据库
    add_kili_pickup_station()
    print("add_kili_pickup_station success")

    # 自提点数据与地址areas汇总
    for it in bee_common_db.Areas.find(
            {"regionCode": {"$in": ["KE", "CN"]}}).sort([("deep", 1)]):
        it["lastUpdatedTime"] = convert_time_str(it["lastUpdatedTime"])
        ki_li_data["areas"].append(it)
    with open("ki_li_data.json", 'w', encoding="utf-8") as write_f:
        write_f.write(json.dumps(ki_li_data, indent=4, ensure_ascii=False))
